import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.worksheet.cell_range import CellRange
# from openpyxl.cell import Cell

try:
    wb:Workbook = Workbook()

    url = "https://m.xpiaotian.com/zh_hant/book/390558/"
    rsp = requests.get(url)

    # 抓取回傳的狀態碼，200表示成功，不是表示有錯誤發生
    if rsp.status_code != 200:
        raise RuntimeError(f'抓取網頁{url}發生錯誤，代碼：{rsp.status_code}')  

    # 也可以用 html5lib，不過要先安裝 pip install html5lib
    soup = BeautifulSoup(rsp.text, 'html.parser')

    # prettify()用來把 html 程式變漂亮
    #print(soup.prettify())

    ws:Worksheet = wb.worksheets[0]
    ws_story:Worksheet = ( wb['內文']
                            if '內文' in wb.sheetnames
                            else wb.create_sheet(title='內文')
    )                       
    ws.title = '目錄'
    ws.append(['集數', '網址'])
    episodes = soup.select('body>ul.chapter:nth-of-type(2)>li')
    # print(episodes)

    for ep in episodes[:-1]:
        ep_name = ep.select_one('a').text.split()[0]
        ep_url = url + ep.select_one('a').attrs["href"].split('/')[-1]
        # print(ep_name, ep_url)
        ws.append([ep_name, ep_url])

        if 1742 <= int(ep_name[1:-1]) <= 1800:
            
            rsp = requests.get(ep_url)
            if rsp.status_code != 200:
                raise RuntimeError(f'抓取網頁{ep_url}發生錯誤，代碼：{rsp.status_code}')  
            
            soup = BeautifulSoup(rsp.text, 'html5lib')
            content = soup.select_one('div#nr1').text.replace(chr(160),'')
            for _ in content.split('\n'): 
                ws_story.append([_])         

except RuntimeError as e:
    print("發生錯誤：",repr(e))  
finally:
    wb.save('小說.xlsx')
    wb.close()